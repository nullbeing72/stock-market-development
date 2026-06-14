"""
nifty100tracker.py — v1.0.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Intelligence features
─────────────────────
① Excel → Model Feedback Loop
   Reads the stock's .xlsx file on every run to extract the history of
   predictions vs actual closes.  Computes per-stock bias, MAPE trend,
   and directional accuracy — then feeds those numbers back into training.

② Adaptive Hyperparameters
   Training config (epochs, patience, direction_weight) auto-adjusted per
   stock based on its rolling accuracy from the Excel log.

③ Drift Detection & Smart Retraining
   Monitors rolling-10-day MAPE vs long-run MAPE.  Triggers full retraining
   when performance degrades beyond threshold; otherwise uses fast incremental
   fine-tuning (CNN weights frozen, head + LSTM + attention updated).

④ Calibrated Predictions
   Applies a signed bias correction (mean historical over/under-prediction)
   so the model self-corrects systematic errors without retraining.

⑤ Model Health Tracking
   JSON sidecar file (<TICKER>_meta.json) stores training date, best val
   loss, current MAPE, directional accuracy, calibration offset, last
   run timestamp — all human-readable at a glance.

⑥ Concurrent Data Fetching
   Market benchmark, bond, and commodity data are fetched concurrently via
   ThreadPoolExecutor so only one blocking network call hits the critical
   path per ticker.

⑦ Data Quality Validation
   OHLCV rows are validated before any downstream use: outlier prices
   (>5σ from 20-day rolling), zero volume, stale dates, and NaN fills
   are flagged, corrected or rejected with clear log messages.

Usage
─────
    python nifty100tracker.py                        # scheduled daemon
    python nifty100tracker.py --once                 # run once, all tickers
    python nifty100tracker.py --once --ticker X.NS   # single stock
    python nifty100tracker.py --once --force         # ignore market-closed
    python nifty100tracker.py --once --retrain       # force full retrain
    python nifty100tracker.py --backfill --days 30   # backfill N days
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed, wait, FIRST_COMPLETED
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import schedule
import torch
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

# ── Ensure data.py and model.py are importable from the same directory ────────
_HERE = Path(__file__).parent
sys.path.insert(0, str(_HERE))

try:
    from data import (
        CLOSE_IDX, FACTOR_FEATURE_COLS, FEATURE_COLS, PRICE_FEATURE_COLS,
        add_quant_features, add_technical_features, fetch_bond_yields,
        fetch_commodities, fetch_data, fetch_market_data,
        build_sequences, split_data, returns_to_prices,
        validate_feature_matrix,                        # ← now imported
    )
    from model import (
        HybridModel, make_loaders, monte_carlo_forecast, predict_test_set,
        train_model, finetune_model, predict_with_calibration, compute_metrics,
    )
    from sklearn.preprocessing import MinMaxScaler
    _ML_AVAILABLE = True
except ImportError as _e:
    _ML_AVAILABLE = False
    print(f"[WARN] ML modules not fully available: {_e}. Predictions will be skipped.")


# ─────────────────────────────────────────────────────────────────────────────
# NIFTY 100 TICKERS  (Yahoo Finance NSE symbols)
# ─────────────────────────────────────────────────────────────────────────────
NIFTY100_TICKERS: list[str] = [
    # Financials
    "HDFCBANK.NS", "ICICIBANK.NS", "KOTAKBANK.NS", "SBIN.NS", "AXISBANK.NS",
    "BAJFINANCE.NS", "BAJAJFINSV.NS", "HDFCLIFE.NS", "SBILIFE.NS", "ICICIGI.NS",
    "INDUSINDBK.NS", "FEDERALBNK.NS", "BANDHANBNK.NS", "PNB.NS", "BANKBARODA.NS",
    "CANBK.NS", "MUTHOOTFIN.NS", "CHOLAFIN.NS", "SHRIRAMFIN.NS",
    # Technology
    "TCS.NS", "INFY.NS", "HCLTECH.NS", "WIPRO.NS", "TECHM.NS",
    "LTIM.NS", "PERSISTENT.NS", "COFORGE.NS", "MPHASIS.NS",
    # Energy & Oil
    "RELIANCE.NS", "ONGC.NS", "BPCL.NS", "IOC.NS", "NTPC.NS",
    "POWERGRID.NS", "ADANIGREEN.NS", "ADANITRANS.NS", "TATAPOWER.NS", "GAIL.NS",
    # Consumer
    "HINDUNILVR.NS", "ITC.NS", "NESTLEIND.NS", "DABUR.NS", "MARICO.NS",
    "GODREJCP.NS", "BRITANNIA.NS", "COLPAL.NS", "EMAMILTD.NS",
    # Industrials & Infra
    "LT.NS", "ADANIENT.NS", "ADANIPORTS.NS", "SIEMENS.NS", "ABB.NS",
    "HAVELLS.NS", "BHEL.NS", "CUMMINSIND.NS", "THERMAX.NS",
    # Automobiles
    "MARUTI.NS", "TATAMOTORS.NS", "M&M.NS", "BAJAJ-AUTO.NS", "EICHERMOT.NS",
    "HEROMOTOCO.NS", "TVSMOTORS.NS", "ASHOKLEY.NS",
    # Healthcare & Pharma
    "SUNPHARMA.NS", "DRREDDY.NS", "CIPLA.NS", "DIVISLAB.NS", "APOLLOHOSP.NS",
    "LUPIN.NS", "TORNTPHARM.NS", "AUROPHARMA.NS", "ALKEM.NS",
    # Metals & Mining
    "TATASTEEL.NS", "JSWSTEEL.NS", "HINDALCO.NS", "VEDL.NS", "COALINDIA.NS",
    "NMDC.NS", "SAIL.NS",
    # Cement
    "ULTRACEMCO.NS", "AMBUJACEM.NS", "ACC.NS", "SHREECEM.NS",
    # Telecom
    "BHARTIARTL.NS", "VODAFONEIDEA.NS",
    # Real Estate
    "DLF.NS", "GODREJPROP.NS", "PRESTIGE.NS",
    # Consumer Discretionary
    "TITAN.NS", "TRENT.NS", "DMART.NS", "NYKAA.NS", "ZOMATO.NS",
    "PAYTM.NS", "INDIAMART.NS",
    # Chemicals
    "PIDILITIND.NS", "ASIANPAINT.NS", "BERGER.NS", "SRF.NS",
    # Logistics
    "IRCTC.NS", "CONCOR.NS",
    # Agriculture
    "UPL.NS", "PIIND.NS",
    # Conglomerates
    "TATACONSUM.NS", "TATACHEM.NS",
]

# Deduplicate while preserving order
_seen: set = set()
NIFTY100_TICKERS = [t for t in NIFTY100_TICKERS if t not in _seen and not _seen.add(t)]

# ─────────────────────────────────────────────────────────────────────────────
# DIRECTORIES & PATHS
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
DATA_DIR   = BASE_DIR / "nifty100_data"
MODELS_DIR = BASE_DIR / "nifty100_models"
LOGS_DIR   = BASE_DIR / "logs"
STATE_FILE = BASE_DIR / "tracker_state.json"

for _d in (DATA_DIR, MODELS_DIR, LOGS_DIR):
    _d.mkdir(parents=True, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
_log_file = LOGS_DIR / f"tracker_{date.today().isoformat()}.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(_log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("nifty100tracker")

# ─────────────────────────────────────────────────────────────────────────────
# NSE MARKET CALENDAR
# ─────────────────────────────────────────────────────────────────────────────
NSE_OPEN_HOUR_IST  = 9
NSE_CLOSE_HOUR_IST = 15

NSE_HOLIDAYS: set[date] = {
    # 2024
    date(2024, 1, 22), date(2024, 1, 26), date(2024, 3, 25), date(2024, 3, 29),
    date(2024, 4, 11), date(2024, 4, 14), date(2024, 4, 17), date(2024, 4, 21),
    date(2024, 5, 23), date(2024, 6, 17), date(2024, 7, 17), date(2024, 8, 15),
    date(2024, 10, 2), date(2024, 10, 12), date(2024, 10, 15), date(2024, 11, 1),
    date(2024, 11, 15), date(2024, 11, 20), date(2024, 12, 25),
    # 2025
    date(2025, 1, 26), date(2025, 2, 26), date(2025, 3, 14), date(2025, 3, 31),
    date(2025, 4, 10), date(2025, 4, 14), date(2025, 4, 18), date(2025, 5, 1),
    date(2025, 6, 7), date(2025, 8, 15), date(2025, 8, 27), date(2025, 10, 2),
    date(2025, 10, 21), date(2025, 10, 22), date(2025, 11, 5),
    date(2025, 11, 12), date(2025, 12, 25),
    # 2026
    date(2026, 1, 26), date(2026, 3, 23), date(2026, 4, 3), date(2026, 4, 10),
    date(2026, 4, 14), date(2026, 8, 15), date(2026, 10, 2), date(2026, 11, 12),
    date(2026, 12, 25),
}

def is_market_open(check_date: Optional[date] = None) -> tuple[bool, str]:
    d = check_date or date.today()
    if d.weekday() >= 5:
        return False, f"{d.strftime('%A')} — NSE closed on weekends"
    if d in NSE_HOLIDAYS:
        return False, f"{d.isoformat()} is an NSE public holiday"
    return True, ""


# ─────────────────────────────────────────────────────────────────────────────
# ① INTELLIGENCE LAYER: Excel History Reader & Analytics
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ErrorStats:
    """Per-stock performance analytics derived from the Excel history."""
    n_rows:             int   = 0       # total rows with valid pred+actual
    mape_all:           float = 0.0     # full-history MAPE (%)
    mape_recent:        float = 0.0     # last 20 rows MAPE (%)
    directional_acc:    float = 0.0     # % of days direction was correct
    bias_offset:        float = 0.0     # signed mean (pred - actual) in ₹
    bias_pct:           float = 0.0     # signed mean % (pred - actual)/actual
    drift_detected:     bool  = False   # True if mape_recent >> mape_all
    std_consistency:    float = 0.0     # mean(std) — high → model uncertain


def _read_excel_history(xlsx_path: Path) -> pd.DataFrame:
    """
    Read all data rows from the Excel file (row 4 onward).
    Returns a DataFrame with at least [Date, Today_Pred_Mean, Actual_Price,
    Error2_Abs, Error2_Pct] when available, plus directional columns.
    Returns empty DataFrame if file missing or unreadable.
    """
    if not xlsx_path.exists():
        return pd.DataFrame()
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        # Row 3 contains column names
        headers = [
            (ws.cell(row=3, column=c).value or f"col_{c}")
            for c in range(1, len(EXCEL_COLUMNS) + 1)
        ]

        rows = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append(dict(zip(headers, row)))
        wb.close()

        if not rows:
            return pd.DataFrame()

        df = pd.DataFrame(rows)

        # Parse Date column robustly
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
            df = df.dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)

        # Coerce numeric columns
        num_cols = [
            "Today_Pred_Mean", "Today_Pred_Std", "Actual_Price",
            "Tomorrow_Pred_Mean", "Tomorrow_Pred_Std",
            "Error2_Abs", "Error2_Pct", "Close",
        ]
        for c in num_cols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")

        return df
    except Exception as e:
        log.warning("Could not read Excel history from %s: %s", xlsx_path.name, e)
        return pd.DataFrame()


def compute_error_stats(hist: pd.DataFrame) -> ErrorStats:
    """
    Derive ErrorStats from the Excel history DataFrame.
    All fields default to neutral / zero when data is insufficient.
    """
    stats = ErrorStats()
    if hist.empty:
        return stats

    # Need at least pred + actual to compute anything
    valid = hist.dropna(subset=["Today_Pred_Mean", "Actual_Price"]).copy()
    stats.n_rows = len(valid)

    if stats.n_rows < 5:
        return stats

    pred   = valid["Today_Pred_Mean"].values
    actual = valid["Actual_Price"].values

    denom  = np.where(np.abs(actual) < 1e-9, 1e-9, actual)
    errors = np.abs((pred - actual) / denom) * 100.0

    stats.mape_all    = float(np.nanmean(errors))
    stats.mape_recent = float(np.nanmean(errors[-20:]))
    stats.bias_offset = float(np.nanmean(pred - actual))
    stats.bias_pct    = float(np.nanmean((pred - actual) / denom) * 100.0)

    # Drift: recent MAPE worse than long-run MAPE by >50%
    if stats.mape_all > 0:
        stats.drift_detected = stats.mape_recent > stats.mape_all * 1.5

    # Directional accuracy from consecutive actual closes
    if len(valid) > 1 and "Actual_Price" in valid.columns:
        a_arr  = valid["Actual_Price"].dropna().values
        p_arr  = valid["Today_Pred_Mean"].values[:len(a_arr)]
        if len(a_arr) > 1:
            correct = np.sign(np.diff(p_arr)) == np.sign(np.diff(a_arr))
            stats.directional_acc = float(np.mean(correct) * 100.0)

    # MC std consistency
    if "Today_Pred_Std" in valid.columns:
        stds = valid["Today_Pred_Std"].dropna()
        if len(stds) > 0:
            stats.std_consistency = float(stds.mean())

    return stats


# ─────────────────────────────────────────────────────────────────────────────
# ② ADAPTIVE HYPERPARAMETERS
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class TrainingConfig:
    """Per-stock adaptive training configuration."""
    epochs:           int   = 30
    patience:         int   = 8
    direction_weight: float = 0.4
    do_finetune:      bool  = False   # True  → incremental fine-tune
    do_retrain:       bool  = True    # False → load from disk, skip training
    reason:           str   = "initial"

    def describe(self) -> str:
        mode = "finetune" if self.do_finetune else ("retrain" if self.do_retrain else "load")
        return (
            f"mode={mode} | epochs={self.epochs} | patience={self.patience} "
            f"| dir_weight={self.direction_weight:.2f} | reason={self.reason}"
        )


def build_training_config(
    ticker:      str,
    stats:       ErrorStats,
    model_path:  Path,
    force_retrain: bool = False,
) -> TrainingConfig:
    """
    Decide what to do with the model based on its performance history.

    Decision tree
    ─────────────
    1. No model file → full initial training.
    2. force_retrain flag set → full retrain.
    3. Drift detected (recent MAPE > 1.5× long-run MAPE) → full retrain.
    4. Model exists + no drift + ≥20 historical rows → incremental fine-tune.
    5. Model exists + too few rows → just load (can't judge performance yet).
    """
    cfg = TrainingConfig()

    # ── Adapt direction weight based on historical accuracy ───────────────
    if stats.n_rows >= 20:
        if stats.directional_acc < 50.0:
            cfg.direction_weight = 0.65   # model is guessing wrong direction
            log.info("[%s] Low directional accuracy (%.1f%%) → direction_weight=0.65",
                     ticker, stats.directional_acc)
        elif stats.directional_acc < 60.0:
            cfg.direction_weight = 0.50
        else:
            cfg.direction_weight = 0.40   # good accuracy; standard weight

    # ── Decide training mode ──────────────────────────────────────────────
    model_exists = model_path.exists()

    if force_retrain:
        cfg.do_retrain  = True
        cfg.do_finetune = False
        cfg.reason      = "forced retrain"
        cfg.epochs      = 40
        return cfg

    if not model_exists:
        cfg.do_retrain  = True
        cfg.do_finetune = False
        cfg.reason      = "no saved model"
        return cfg

    if stats.n_rows < 20:
        cfg.do_retrain  = False
        cfg.do_finetune = False
        cfg.reason      = "insufficient history — load only"
        return cfg

    if stats.drift_detected:
        cfg.do_retrain  = True
        cfg.do_finetune = False
        cfg.reason      = (
            f"drift detected: recent MAPE={stats.mape_recent:.2f}% "
            f"vs long-run={stats.mape_all:.2f}%"
        )
        cfg.epochs  = 35
        cfg.patience = 10
        log.warning("[%s] Performance drift detected → full retrain", ticker)
        return cfg

    # Model exists, no drift, enough history → fine-tune
    cfg.do_retrain  = False
    cfg.do_finetune = True
    cfg.epochs      = 8
    cfg.patience    = 4
    cfg.reason      = (
        f"incremental fine-tune: MAPE={stats.mape_recent:.2f}% "
        f"dir={stats.directional_acc:.1f}%"
    )
    return cfg


# ─────────────────────────────────────────────────────────────────────────────
# ⑤ MODEL HEALTH TRACKING (JSON sidecar)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ModelMeta:
    ticker:            str   = ""
    trained_on:        str   = ""    # ISO date of last training/fine-tune
    training_mode:     str   = ""    # "full" | "finetune" | "loaded"
    best_val_loss:     float = 0.0
    epochs_run:        int   = 0
    mape_at_train:     float = 0.0
    directional_acc:   float = 0.0
    bias_offset:       float = 0.0   # ₹ correction applied
    calibration_on:    bool  = False
    n_history_rows:    int   = 0
    drift_detected:    bool  = False
    last_run:          str   = ""
    run_count:         int   = 0
    notes:             str   = ""


def _meta_path(ticker: str) -> Path:
    safe = ticker.replace(".NS", "").replace("/", "_")
    return MODELS_DIR / f"{safe}_meta.json"


def load_model_meta(ticker: str) -> ModelMeta:
    path = _meta_path(ticker)
    if path.exists():
        try:
            data = json.loads(path.read_text())
            return ModelMeta(**{k: v for k, v in data.items() if k in ModelMeta.__dataclass_fields__})
        except Exception:
            pass
    return ModelMeta(ticker=ticker)


def save_model_meta(meta: ModelMeta) -> None:
    path = _meta_path(meta.ticker)
    path.write_text(json.dumps(asdict(meta), indent=2, default=str))


# ─────────────────────────────────────────────────────────────────────────────
# STATE MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except Exception:
            pass
    return {}


def save_state(state: dict) -> None:
    STATE_FILE.write_text(json.dumps(state, indent=2, default=str))


def was_processed_today(state: dict, ticker: str) -> bool:
    last = state.get(ticker, {}).get("last_run")
    return last == date.today().isoformat()


def mark_processed(state: dict, ticker: str, status: str = "ok") -> None:
    state[ticker] = {
        "last_run":  date.today().isoformat(),
        "status":    status,
        "timestamp": datetime.now().isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# ⑦ DATA QUALITY VALIDATION
# ─────────────────────────────────────────────────────────────────────────────

def validate_ohlcv(df: pd.DataFrame, ticker: str) -> tuple[pd.DataFrame, list[str]]:
    """
    Validate and clean OHLCV data. Returns (cleaned_df, list_of_warnings).

    Checks:
    • High >= Low, High >= Close, Low <= Open
    • Price outliers: flag rows where Close is >5σ from 20-day rolling mean
    • Volume = 0 on trading days (replaced with forward-fill)
    • Duplicate index dates
    • Insufficient rows
    """
    warnings: list[str] = []

    if df is None or df.empty:
        return df, ["No data returned"]

    # Deduplicate index
    dupes = df.index.duplicated()
    if dupes.any():
        warnings.append(f"{dupes.sum()} duplicate dates removed")
        df = df[~dupes]

    # Ensure required columns
    required = {"Open", "High", "Low", "Close", "Volume"}
    missing  = required - set(df.columns)
    if missing:
        warnings.append(f"Missing columns: {missing}")
        return df, warnings

    # OHLC sanity: H >= L, H >= C, L <= O
    bad_hl = (df["High"] < df["Low"]).sum()
    if bad_hl:
        warnings.append(f"{bad_hl} rows with High < Low; correcting via swap")
        mask = df["High"] < df["Low"]
        df.loc[mask, ["High", "Low"]] = df.loc[mask, ["Low", "High"]].values

    # Outlier detection: |Close - rolling_mean| > 5σ
    rolling_mean = df["Close"].rolling(20, min_periods=5).mean()
    rolling_std  = df["Close"].rolling(20, min_periods=5).std()
    outliers = (df["Close"] - rolling_mean).abs() > 5 * rolling_std.fillna(1e9)
    if outliers.any():
        n_out = outliers.sum()
        warnings.append(f"{n_out} price outlier(s) detected (>5σ) — replaced with ffill")
        df.loc[outliers, "Close"] = np.nan
        df["Close"] = df["Close"].ffill()

    # Zero volume on trading days
    zero_vol = (df["Volume"] == 0)
    if zero_vol.any():
        warnings.append(f"{zero_vol.sum()} rows with zero volume — forward-filled")
        df.loc[zero_vol, "Volume"] = np.nan
        df["Volume"] = df["Volume"].ffill().fillna(0)

    # Stale data check: last row shouldn't be older than 5 trading days
    if len(df) > 0:
        latest = df.index[-1]
        if hasattr(latest, "date"):
            latest = latest.date()
        lag = (date.today() - latest).days
        if lag > 7:
            warnings.append(f"Data appears stale: last row is {lag} days ago ({latest})")

    if len(df) < 80:
        warnings.append(f"Only {len(df)} rows — ML predictions will be skipped")

    return df, warnings


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL TEMPLATE & IO
# ─────────────────────────────────────────────────────────────────────────────

EXCEL_COLUMNS = [
    "Date",
    "Open", "High", "Low", "Close", "Volume",
    "SMA_10", "SMA_50", "MACD", "MACD_Signal", "RSI", "ROC_10",
    "Stoch_K", "BB_Upper", "BB_Lower", "ATR", "OBV_norm", "RVI",
    "Beta_60", "Alpha_60", "Momentum_12_1", "Sharpe_60", "Vol_Ratio", "Mkt_Return",
    "Yest_Pred_Mean", "Yest_Pred_Std",
    "Today_Pred_Mean", "Today_Pred_Std",
    "Actual_Price",
    "Tomorrow_Pred_Mean", "Tomorrow_Pred_Std",
    "Error1_Abs", "Error1_Pct",
    "Error2_Abs", "Error2_Pct",
]

_SECTION_HEADERS = [
    (1,  1,  "DATE"),
    (2,  6,  "OHLCV"),
    (7,  24, "ENGINEERED FEATURES"),
    (25, 26, "YESTERDAY'S PREDICTION"),
    (27, 28, "TODAY'S PREDICTION"),
    (29, 29, "ACTUAL PRICE"),
    (30, 31, "TOMORROW'S PREDICTION"),
    (32, 33, "ERROR 1 (|YEST PRED – TODAY PRED|)"),
    (34, 35, "ERROR 2 (|TODAY PRED – ACTUAL PRICE|)"),
]

_STYLE_HEADER_TOP = {
    "font":      Font(bold=True, color="FFFFFF", size=10),
    "fill":      PatternFill("solid", fgColor="1F3864"),
    "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
}
_STYLE_HEADER_MID = {
    "font":      Font(bold=True, size=9),
    "fill":      PatternFill("solid", fgColor="BDD7EE"),
    "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
}
_STYLE_DATA = {
    "font":      Font(size=9),
    "alignment": Alignment(horizontal="center", vertical="center"),
}
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _apply_style(cell, style_dict: dict) -> None:
    for attr, val in style_dict.items():
        setattr(cell, attr, val)
    cell.border = _THIN_BORDER


def create_xlsx_template(path: Path, company_name: str, ticker: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = ticker.replace(".NS", "")

    ws.merge_cells(f"A1:{get_column_letter(len(EXCEL_COLUMNS))}1")
    ws["A1"] = f"COMPANY: {company_name}  |  TICKER: {ticker}"
    ws["A1"].font      = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c_start, c_end, label in _SECTION_HEADERS:
        col_s = get_column_letter(c_start)
        col_e = get_column_letter(c_end)
        cell_ref = f"{col_s}2"
        if c_start != c_end:
            ws.merge_cells(f"{col_s}2:{col_e}2")
        ws[cell_ref] = label
        _apply_style(ws[cell_ref], _STYLE_HEADER_TOP)
    ws.row_dimensions[2].height = 36

    for i, col in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=3, column=i, value=col)
        _apply_style(cell, _STYLE_HEADER_MID)
        ws.column_dimensions[get_column_letter(i)].width = (
            12 if i == 1 else 8 if i <= 6 else 11 if i <= 24 else 13
        )
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "A4"

    wb.save(path)
    log.info("Created Excel template: %s", path.name)


def _get_company_xlsx(ticker: str, company_name: str) -> Path:
    safe_name = ticker.replace(".NS", "").replace("/", "_").replace("\\", "_")
    path = DATA_DIR / f"{safe_name}.xlsx"
    if not path.exists():
        create_xlsx_template(path, company_name, ticker)
    return path


def _existing_dates_in_xlsx(path: Path) -> set[str]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        dates: set[str] = set()
        for row in ws.iter_rows(min_row=4, max_col=1, values_only=True):
            v = row[0]
            if v is None:
                continue
            if isinstance(v, (date, datetime)):
                dates.add(v.strftime("%Y-%m-%d"))
            else:
                s = str(v)
                # Try both DD-MM-YYYY and ISO format
                try:
                    dates.add(pd.to_datetime(s, dayfirst=True).strftime("%Y-%m-%d"))
                except Exception:
                    dates.add(s[:10])
        wb.close()
        return dates
    except Exception as e:
        log.warning("Could not read existing dates from %s: %s", path.name, e)
        return set()


def append_row_to_xlsx(path: Path, row_data: dict) -> bool:
    """
    Append a single data row. Returns True on success.

    Error column font colouring (v1.0.0):
    ──────────────────────────────────────
    Values stored are always absolute (|pred − ref|) so the model has a
    clean magnitude signal.  The *font colour* carries direction:

      Error 2 (TodayPred vs Actual):
        GREEN  → model under-predicted  (pred < actual, actual beat expectations)
        RED    → model over-predicted   (pred > actual, model was too optimistic)

      Error 1 (YestPred vs TodayPred):
        GREEN  → model revised prediction DOWN (became more conservative)
        RED    → model revised prediction UP   (became more aggressive)

    A neutral/missing signed value keeps the default text colour.
    """
    # ── Derive signed values for direction colouring ──────────────────────
    today_pred  = row_data.get("Today_Pred_Mean")
    actual      = row_data.get("Actual_Price")
    yest_pred   = row_data.get("Yest_Pred_Mean")

    # Positive  → over-prediction  → RED    (font "FF4136")
    # Negative  → under-prediction → GREEN  (font "2ECC71")
    e2_signed: Optional[float] = None
    if today_pred is not None and actual is not None:
        try:
            e2_signed = float(today_pred) - float(actual)
        except (TypeError, ValueError):
            pass

    e1_signed: Optional[float] = None
    if yest_pred is not None and today_pred is not None:
        try:
            e1_signed = float(yest_pred) - float(today_pred)
        except (TypeError, ValueError):
            pass

    _FONT_RED   = Font(size=9, color="D92B2B", bold=True)   # over-prediction / revision up
    _FONT_GREEN = Font(size=9, color="1A8F4C", bold=True)   # under-prediction / revision down
    _FONT_DEF   = Font(size=9)                               # no signal available

    def _error_font(signed: Optional[float]) -> Font:
        if signed is None:
            return _FONT_DEF
        return _FONT_RED if signed > 0 else _FONT_GREEN

    # Map column name → font to use (only error columns get special treatment)
    _error_font_map: dict[str, Font] = {
        "Error1_Abs": _error_font(e1_signed),
        "Error1_Pct": _error_font(e1_signed),
        "Error2_Abs": _error_font(e2_signed),
        "Error2_Pct": _error_font(e2_signed),
    }

    try:
        wb = load_workbook(path)
        ws = wb.active
        next_row = ws.max_row + 1

        row_values = [row_data.get(col) for col in EXCEL_COLUMNS]
        for col_idx, (col_name, value) in enumerate(
            zip(EXCEL_COLUMNS, row_values), start=1
        ):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            _apply_style(cell, _STYLE_DATA)

            # Override font for error columns
            if col_name in _error_font_map:
                cell.font = _error_font_map[col_name]

            if col_idx == 1:
                cell.number_format = "DD-MM-YYYY"
            elif col_idx in range(2, 7):
                cell.number_format = "#,##0.00" if col_idx < 6 else "#,##0"
            elif col_idx >= 25:
                cell.number_format = "0.00"

        if next_row % 2 == 0:
            fill = PatternFill("solid", fgColor="EBF3FB")
            for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
                ws.cell(row=next_row, column=col_idx).fill = fill

        wb.save(path)
        return True
    except Exception as e:
        log.error("Failed to append row to %s: %s", path.name, e)
        return False


# ─────────────────────────────────────────────────────────────────────────────
# MODEL MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

SEQ_LENGTH    = 60
PERIOD_DAYS   = 700
FORECAST_DAYS = 2
DEVICE        = torch.device("cuda" if torch.cuda.is_available() else "cpu")

_MODEL_HYPERPARAMS = dict(
    hidden_size=128, num_layers=2,
    cnn_channels=64, num_heads=4, dropout=0.2,
)


def _model_path(ticker: str) -> Path:
    safe = ticker.replace(".NS", "").replace("/", "_")
    return MODELS_DIR / f"{safe}_model.pt"


def _load_model(ticker: str, input_size: int) -> Optional[HybridModel]:
    mpath = _model_path(ticker)
    if not mpath.exists():
        return None
    try:
        model = HybridModel(input_size=input_size, **_MODEL_HYPERPARAMS).to(DEVICE)
        model.load_state_dict(torch.load(mpath, map_location=DEVICE, weights_only=True))
        log.info("[%s] Loaded saved model weights", ticker)
        return model
    except Exception as e:
        log.warning("[%s] Could not load model (%s); will retrain", ticker, e)
        return None


def _save_model(model: HybridModel, ticker: str) -> None:
    mpath = _model_path(ticker)
    torch.save(model.state_dict(), mpath)
    log.info("[%s] Model weights saved → %s", ticker, mpath.name)


def _get_model(
    ticker:       str,
    cfg:          TrainingConfig,
    X_train, y_train,
    X_val,   y_val,
    X_test,  y_test,
    input_size:   int,
) -> tuple[Optional[HybridModel], str, float, int]:
    """
    Returns (model, training_mode, best_val_loss, epochs_run).
    Applies the TrainingConfig to decide between load / fine-tune / retrain.
    """
    train_loader, val_loader, test_loader = make_loaders(
        X_train, y_train, X_val, y_val, X_test, y_test, batch_size=32,
    )

    # ── Load-only ────────────────────────────────────────────────────────
    if not cfg.do_retrain and not cfg.do_finetune:
        model = _load_model(ticker, input_size)
        if model is None:
            log.warning("[%s] No model on disk; falling back to full retrain", ticker)
        else:
            return model, "loaded", 0.0, 0

    # ── Incremental fine-tuning ───────────────────────────────────────────
    if cfg.do_finetune and not cfg.do_retrain:
        model = _load_model(ticker, input_size)
        if model is None:
            log.warning("[%s] No model to fine-tune; falling back to full retrain", ticker)
        else:
            log.info("[%s] Fine-tuning: epochs=%d patience=%d dir_weight=%.2f",
                     ticker, cfg.epochs, cfg.patience, cfg.direction_weight)
            model, history = finetune_model(
                model, train_loader, val_loader,
                device=DEVICE,
                epochs=cfg.epochs,
                patience=cfg.patience,
                direction_weight=cfg.direction_weight,
            )
            best_val = min(history["val"]) if history["val"] else 0.0
            _save_model(model, ticker)
            return model, "finetune", best_val, len(history["val"])

    # ── Full retrain ──────────────────────────────────────────────────────
    log.info("[%s] Full training: epochs=%d patience=%d dir_weight=%.2f",
             ticker, cfg.epochs, cfg.patience, cfg.direction_weight)
    try:
        model, history = train_model(
            train_loader, val_loader,
            input_size=input_size,
            epochs=cfg.epochs,
            patience=cfg.patience,
            device=DEVICE,
            direction_weight=cfg.direction_weight,
            **_MODEL_HYPERPARAMS,
        )
        best_val = min(history["val"]) if history["val"] else 0.0
        _save_model(model, ticker)
        log.info("[%s] Training done. Best val loss: %.6f | Epochs run: %d",
                 ticker, best_val, len(history["val"]))
        return model, "retrain", best_val, len(history["val"])
    except Exception as e:
        log.error("[%s] Model training failed: %s", ticker, e)
        return None, "error", 0.0, 0


# ─────────────────────────────────────────────────────────────────────────────
# ⑥ CONCURRENT DATA FETCHING
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_auxiliary_data(period_days: int) -> tuple:
    """
    Fetch market benchmark, bond yields, and commodities concurrently.
    Returns (market_df, bond_df, comm_df) — any may be None on failure.
    """
    results: dict = {}

    def _do(key, fn, *args):
        try:
            results[key] = fn(*args)
        except Exception as e:
            log.warning("Concurrent fetch failed for '%s': %s", key, e)
            results[key] = None

    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = {
            pool.submit(_do, "market", fetch_market_data, period_days): "market",
            pool.submit(_do, "bonds",  fetch_bond_yields,  period_days): "bonds",
            pool.submit(_do, "comms",  fetch_commodities,  period_days): "comms",
        }
        for f in as_completed(futures):
            pass  # results populated via _do side-effect

    return results.get("market"), results.get("bonds"), results.get("comms")


# ─────────────────────────────────────────────────────────────────────────────
# RETRY DECORATOR
# ─────────────────────────────────────────────────────────────────────────────

def _retry(fn, *args, retries: int = 3, delay: float = 5.0, **kwargs):
    for attempt in range(1, retries + 1):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            if attempt == retries:
                raise
            wait = delay * (2 ** (attempt - 1))
            log.warning("Attempt %d/%d failed (%s). Retrying in %.0fs…", attempt, retries, e, wait)
            time.sleep(wait)


def _get_yesterday_prediction(xlsx_path: Path) -> tuple[Optional[float], Optional[float]]:
    try:
        wb  = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws  = wb.active
        if ws.max_row < 4:
            wb.close()
            return None, None
        mean_col = EXCEL_COLUMNS.index("Tomorrow_Pred_Mean") + 1
        std_col  = EXCEL_COLUMNS.index("Tomorrow_Pred_Std")  + 1
        last_row = ws.max_row
        mean_val = ws.cell(row=last_row, column=mean_col).value
        std_val  = ws.cell(row=last_row, column=std_col).value
        wb.close()
        return (
            float(mean_val) if mean_val is not None else None,
            float(std_val)  if std_val  is not None else None,
        )
    except Exception:
        return None, None


# ─────────────────────────────────────────────────────────────────────────────
# CORE PROCESSING PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def process_ticker(
    ticker:        str,
    company_name:  str,
    run_date:      Optional[date] = None,
    force_retrain: bool = False,
) -> str:
    """
    Full intelligence pipeline for a single ticker.
    Returns 'ok', 'skipped_duplicate', or 'error'.
    """
    target_date = run_date or date.today()
    log.info("─" * 65)
    log.info("[%s] Starting pipeline for %s", ticker, target_date.isoformat())

    # ── 1. Excel dedup check & history read ──────────────────────────────
    xlsx_path = _get_company_xlsx(ticker, company_name)
    existing  = _existing_dates_in_xlsx(xlsx_path)
    date_key  = target_date.isoformat()

    if date_key in existing:
        log.info("[%s] %s already in file — skipping", ticker, date_key)
        return "skipped_duplicate"

    # ── INTELLIGENCE: Read Excel history & compute error stats ────────────
    hist         = _read_excel_history(xlsx_path)
    error_stats  = compute_error_stats(hist)
    if error_stats.n_rows > 0:
        log.info(
            "[%s] Excel history: %d rows | MAPE=%.2f%% | MAPE-recent=%.2f%% "
            "| Dir-acc=%.1f%% | Bias=₹%.2f | Drift=%s",
            ticker,
            error_stats.n_rows,
            error_stats.mape_all,
            error_stats.mape_recent,
            error_stats.directional_acc,
            error_stats.bias_offset,
            "YES ⚠" if error_stats.drift_detected else "no",
        )

    # ── 2. Fetch OHLCV (with retry) ───────────────────────────────────────
    log.info("[%s] Fetching price data…", ticker)
    raw_df = _retry(fetch_data, ticker, PERIOD_DAYS, retries=3, delay=5.0)
    if raw_df is None or raw_df.empty:
        log.error("[%s] No price data returned", ticker)
        return "error"

    # ── INTELLIGENCE: Validate & clean data ───────────────────────────────
    raw_df, val_warnings = validate_ohlcv(raw_df, ticker)
    for w in val_warnings:
        log.warning("[%s] Data validation: %s", ticker, w)

    # ── 3. Feature engineering (concurrent auxiliary fetch) ───────────────
    log.info("[%s] Engineering features (parallel aux fetch)…", ticker)
    try:
        mkt, bonds, comms = _fetch_auxiliary_data(PERIOD_DAYS)

        if mkt is None or mkt.empty:
            log.warning("[%s] Market data unavailable — using stock data as proxy", ticker)
            mkt = raw_df

        df = add_technical_features(raw_df)
        df = add_quant_features(df, mkt, bonds, comms)
        df = df.dropna(subset=["Close"])
    except Exception as e:
        log.error("[%s] Feature engineering failed: %s\n%s", ticker, e, traceback.format_exc())
        return "error"

    # Target date row (or fallback to latest)
    if pd.Timestamp(target_date) in df.index:
        today_row = df.loc[pd.Timestamp(target_date)]
    elif len(df) > 0:
        today_row = df.iloc[-1]
    else:
        log.error("[%s] No rows after feature engineering", ticker)
        return "error"

    actual_price = float(today_row.get("Close", np.nan))

    # ── 4. ML Predictions ─────────────────────────────────────────────────
    today_pred_mean = today_pred_std = None
    tom_pred_mean   = tom_pred_std   = None
    yest_pred_mean  = yest_pred_std  = None
    calibration_applied = False

    if _ML_AVAILABLE:
        try:
            # ── FIX: validate & sanitise feature matrix before scaling ────
            # validate_feature_matrix fills missing columns with 0, zeroes
            # all-NaN columns (e.g. Bond_5Y/Bond_10Y when fetch fails, or
            # commodity columns), and replaces ±Inf.  Without this step,
            # MinMaxScaler propagates NaN through the entire scaled array,
            # build_sequences produces all-NaN targets, and training /
            # inference silently outputs NaN → predictions stay None.
            df_validated, feat_warnings = validate_feature_matrix(
                df, FEATURE_COLS, ticker=ticker,
            )
            for w in feat_warnings:
                log.warning("[%s] Feature validation: %s", ticker, w)

            feature_df = df_validated[FEATURE_COLS].copy()
            raw_close  = df_validated["Close"].values.astype(np.float32)

            n_rows = len(feature_df)
            log.info("[%s] Feature matrix: %d rows × %d cols", ticker, n_rows, len(FEATURE_COLS))

            if n_rows >= SEQ_LENGTH + 20:
                scaler = MinMaxScaler()
                scaled = scaler.fit_transform(feature_df.values).astype(np.float32)

                # Guard: if scaler produced NaN (e.g. zero-range column
                # that validate_feature_matrix missed), abort ML gracefully
                nan_count = int(np.isnan(scaled).sum())
                if nan_count > 0:
                    log.warning(
                        "[%s] Scaled feature matrix contains %d NaN values after "
                        "MinMaxScaler — replacing with 0 before sequencing",
                        ticker, nan_count,
                    )
                    scaled = np.nan_to_num(scaled, nan=0.0, posinf=0.0, neginf=0.0)

                X_all, y_all = build_sequences(scaled, raw_close, SEQ_LENGTH)

                if len(X_all) == 0:
                    log.warning("[%s] build_sequences returned empty arrays — skipping ML", ticker)
                else:
                    (X_tr, y_tr), (X_val, y_val), (X_te, y_te), val_end, tr_end = split_data(
                        X_all, y_all
                    )
                    input_size = X_tr.shape[2]

                    # ── INTELLIGENCE: Adaptive training config ────────────
                    train_cfg = build_training_config(
                        ticker, error_stats, _model_path(ticker), force_retrain,
                    )
                    log.info("[%s] Training config: %s", ticker, train_cfg.describe())

                    model, train_mode, best_val, epochs_run = _get_model(
                        ticker, train_cfg,
                        X_tr, y_tr, X_val, y_val, X_te, y_te,
                        input_size,
                    )

                    if model is not None:
                        raw_forecasts = monte_carlo_forecast(
                            model=model,
                            scaled_data=scaled,
                            raw_close_last=actual_price,
                            scaler=scaler,
                            input_size=input_size,
                            seq_length=SEQ_LENGTH,
                            forecast_days=FORECAST_DAYS,
                            device=DEVICE,
                            n_samples=80,
                            close_idx=CLOSE_IDX,
                        )

                        # ── INTELLIGENCE: Bias calibration ────────────────
                        calibrated = raw_forecasts
                        if abs(error_stats.bias_offset) > 0.5 and error_stats.n_rows >= 10:
                            calibrated = predict_with_calibration(
                                raw_forecasts,
                                bias_offset=error_stats.bias_offset,
                                bias_pct=error_stats.bias_pct,
                            )
                            calibration_applied = True
                            log.info(
                                "[%s] Calibration applied: bias_offset=₹%.2f",
                                ticker, error_stats.bias_offset,
                            )

                        if len(calibrated) >= 2:
                            today_pred_mean = round(calibrated[0]["mean"], 4)
                            today_pred_std  = round(calibrated[0]["std"],  4)
                            tom_pred_mean   = round(calibrated[1]["mean"], 4)
                            tom_pred_std    = round(calibrated[1]["std"],  4)
                            log.info(
                                "[%s] Forecast OK — Today=%.4f±%.4f  Tomorrow=%.4f±%.4f",
                                ticker,
                                today_pred_mean, today_pred_std,
                                tom_pred_mean,   tom_pred_std,
                            )
                        else:
                            log.warning(
                                "[%s] monte_carlo_forecast returned fewer than 2 steps "
                                "(got %d) — predictions will be blank",
                                ticker, len(calibrated),
                            )

                        yest_pred_mean, yest_pred_std = _get_yesterday_prediction(xlsx_path)

                        # ── INTELLIGENCE: Update model metadata ───────────
                        meta = load_model_meta(ticker)
                        meta.ticker          = ticker
                        meta.trained_on      = date.today().isoformat()
                        meta.training_mode   = train_mode
                        meta.best_val_loss   = round(best_val, 6)
                        meta.epochs_run      = epochs_run
                        meta.mape_at_train   = round(error_stats.mape_recent, 3)
                        meta.directional_acc = round(error_stats.directional_acc, 2)
                        meta.bias_offset     = round(error_stats.bias_offset, 4)
                        meta.calibration_on  = calibration_applied
                        meta.n_history_rows  = error_stats.n_rows
                        meta.drift_detected  = error_stats.drift_detected
                        meta.last_run        = datetime.now().isoformat()
                        meta.run_count       = (meta.run_count or 0) + 1
                        save_model_meta(meta)

            else:
                log.warning("[%s] Not enough rows (%d) for ML — need ≥%d.",
                            ticker, n_rows, SEQ_LENGTH + 20)

        except Exception as e:
            log.warning(
                "[%s] ML pipeline failed: %s\n%s",
                ticker, e, traceback.format_exc(),
            )

    # ── 5. Error metrics ──────────────────────────────────────────────────
    error1_abs = error1_pct = error2_abs = error2_pct = None

    if yest_pred_mean is not None and today_pred_mean is not None:
        error1_abs = abs(yest_pred_mean - today_pred_mean)
        if today_pred_mean and today_pred_mean != 0:
            error1_pct = (error1_abs / abs(today_pred_mean)) * 100.0

    if today_pred_mean is not None and not np.isnan(actual_price):
        error2_abs = abs(today_pred_mean - actual_price)
        if actual_price and actual_price != 0:
            error2_pct = (error2_abs / abs(actual_price)) * 100.0

    # ── 6. Build row dict ─────────────────────────────────────────────────
    def _safe(val):
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return None
        return round(float(val), 6) if isinstance(val, (float, np.floating)) else val

    row: dict = {
        "Date":   target_date.strftime("%d-%m-%Y"),
        "Open":   _safe(today_row.get("Open")),
        "High":   _safe(today_row.get("High")),
        "Low":    _safe(today_row.get("Low")),
        "Close":  _safe(today_row.get("Close")),
        "Volume": int(today_row.get("Volume", 0) or 0),
    }
    for col in ["SMA_10", "SMA_50", "MACD", "MACD_Signal", "RSI", "ROC_10",
                "Stoch_K", "BB_Upper", "BB_Lower", "ATR", "OBV_norm", "RVI"]:
        row[col] = _safe(today_row.get(col))

    for col in ["Beta_60", "Alpha_60", "Momentum_12_1", "Sharpe_60",
                "Vol_Ratio", "Mkt_Return"]:
        row[col] = _safe(today_row.get(col))

    row["Yest_Pred_Mean"]     = yest_pred_mean
    row["Yest_Pred_Std"]      = yest_pred_std
    row["Today_Pred_Mean"]    = today_pred_mean
    row["Today_Pred_Std"]     = today_pred_std
    row["Actual_Price"]       = _safe(actual_price)
    row["Tomorrow_Pred_Mean"] = tom_pred_mean
    row["Tomorrow_Pred_Std"]  = tom_pred_std
    row["Error1_Abs"] = round(error1_abs, 4) if error1_abs is not None else None
    row["Error1_Pct"] = round(error1_pct, 4) if error1_pct is not None else None
    row["Error2_Abs"] = round(error2_abs, 4) if error2_abs is not None else None
    row["Error2_Pct"] = round(error2_pct, 4) if error2_pct is not None else None

    # ── 7. Append to Excel ────────────────────────────────────────────────
    ok = append_row_to_xlsx(xlsx_path, row)
    if ok:
        log.info(
            "[%s] ✓ Row appended | Close=%.2f | TodayPred=%.2f | TomPred=%.2f | "
            "Err2=%.2f%% | Calib=%s",
            ticker,
            actual_price,
            today_pred_mean or 0,
            tom_pred_mean   or 0,
            error2_pct      or 0,
            "✓" if calibration_applied else "—",
        )
        return "ok"
    return "error"


# ─────────────────────────────────────────────────────────────────────────────
# COMPANY NAME LOOKUP
# ─────────────────────────────────────────────────────────────────────────────

_NAME_CACHE: dict[str, str] = {}


def get_company_name(ticker: str) -> str:
    if ticker in _NAME_CACHE:
        return _NAME_CACHE[ticker]
    try:
        import yfinance as yf
        info = yf.Ticker(ticker).info
        name = info.get("longName") or info.get("shortName") or ticker
    except Exception:
        name = ticker.replace(".NS", "")
    _NAME_CACHE[ticker] = name
    return name


# ─────────────────────────────────────────────────────────────────────────────
# BATCH RUN
# ─────────────────────────────────────────────────────────────────────────────

def run_all(
    tickers:       list[str],
    run_date:      Optional[date] = None,
    force:         bool = False,
    force_retrain: bool = False,
) -> dict[str, str]:
    target = run_date or date.today()
    state  = load_state()
    results: dict[str, str] = {}

    log.info("══════════════════════════════════════════════════════════════")
    log.info("NIFTY 100 Tracker  v1.0.0")
    log.info("Run date: %s | Tickers: %d | Device: %s", target.isoformat(), len(tickers), DEVICE)
    log.info("══════════════════════════════════════════════════════════════")

    for i, ticker in enumerate(tickers, start=1):
        log.info("[%d/%d] Processing: %s", i, len(tickers), ticker)

        if not force and was_processed_today(state, ticker):
            log.info("[%s] Already processed today (use --force to override)", ticker)
            results[ticker] = "skipped_already_done"
            continue

        try:
            company = get_company_name(ticker)
            status  = process_ticker(
                ticker, company,
                run_date=target,
                force_retrain=force_retrain,
            )
        except Exception as e:
            log.error("[%s] Unhandled error: %s\n%s", ticker, e, traceback.format_exc())
            status = "error"

        results[ticker] = status
        mark_processed(state, ticker, status)
        save_state(state)
        time.sleep(1.5)   # avoid Yahoo Finance rate-limiting

    ok      = sum(1 for s in results.values() if s == "ok")
    skipped = sum(1 for s in results.values() if "skipped" in s)
    errors  = sum(1 for s in results.values() if s == "error")
    log.info("══ SUMMARY: %d OK | %d Skipped | %d Errors ══", ok, skipped, errors)
    return results


def scheduled_run() -> None:
    open_, reason = is_market_open()
    if not open_:
        log.info("Market closed: %s — skipping run", reason)
        return
    run_all(NIFTY100_TICKERS)


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="NIFTY 100 Intelligent Stock Tracker v1.0.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--once",          action="store_true",
                        help="Run once and exit")
    parser.add_argument("--force",         action="store_true",
                        help="Skip market-open check and duplicate detection")
    parser.add_argument("--retrain",       action="store_true",
                        help="Force full model retraining for each ticker")
    parser.add_argument("--ticker",        type=str, default=None,
                        help="Process only this ticker, e.g. RELIANCE.NS")
    parser.add_argument("--backfill",      action="store_true",
                        help="Backfill historical trading days")
    parser.add_argument("--days",          type=int, default=30,
                        help="Number of days to backfill (default: 30)")
    parser.add_argument("--schedule-time", type=str, default="16:00",
                        help="Daily run time in IST 24h (default: 16:00)")
    parser.add_argument("--show-meta",     type=str, default=None, metavar="TICKER",
                        help="Print model health JSON for a ticker and exit")
    args = parser.parse_args()

    # Show model metadata and exit
    if args.show_meta:
        t = args.show_meta.upper()
        if not t.endswith(".NS"):
            t += ".NS"
        meta = load_model_meta(t)
        print(json.dumps(asdict(meta), indent=2, default=str))
        return

    tickers = [args.ticker] if args.ticker else NIFTY100_TICKERS

    # Backfill mode
    if args.backfill:
        log.info("Backfill mode: last %d trading days", args.days)
        end   = date.today()
        start = end - timedelta(days=args.days)
        for bday in pd.bdate_range(start, end):
            d = bday.date()
            open_, reason = is_market_open(d)
            if not open_:
                log.info("Skipping %s: %s", d.isoformat(), reason)
                continue
            log.info("── Backfilling %s ──", d.isoformat())
            run_all(tickers, run_date=d, force=True, force_retrain=False)
        return

    # Once mode
    if args.once:
        open_, reason = is_market_open()
        if not args.force and not open_:
            log.info("Market closed: %s", reason)
            log.info("Use --force to run anyway")
            return
        run_all(tickers, force=args.force, force_retrain=args.retrain)
        return

    # Scheduled daemon
    log.info("Scheduler started — daily run at %s IST", args.schedule_time)
    schedule.every().day.at(args.schedule_time).do(scheduled_run)
    while True:
        schedule.run_pending()
        time.sleep(30)


if __name__ == "__main__":
    main()
